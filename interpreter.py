import pandas as pd
import os
import webbrowser
import subprocess
from xlsx2html import xlsx2html
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from antlr4 import *
from QAReporterDSLParser import QAReporterDSLParser

# Color constants (hexadecimal)
COLORS_HEX = {
    'red':    'FFFF0000',
    'green':  'FF00FF00',
    'yellow': 'FFFFFF00',
    'orange': 'FFA500',
}

class Interpeter:
    def __init__(self):
        self.df = pd.DataFrame()
        self.styles = {}
        self.filter = [] 
        self.output_file = None

    def run(self, tree):
        self.evaluate(tree)

    # Program execution starts here
    def evaluate(self, t):

        if t is None: 
            return None
        
        # Navigation through Abstract Syntax Tree (AST)
        match t:
            case QAReporterDSLParser.ProgContext():
                for cmd in t.cmd():
                    self.evaluate(cmd)
                return
            
            case QAReporterDSLParser.ReadContext():
                path = t.INPUTFILE().getText().strip('"')
                try:
                    self.df = pd.read_csv(path)
                    print(f"[READ] File '{path}' loaded ({len(self.df)} lines).")
                except Exception as e:
                    print(f"[ERROR] Failed to read CSV file: {e}")
                return

            case QAReporterDSLParser.StyleContext():
                name = t.ID().getText()
                props = {}
                for p in t.props().prop():
                    key = p.key().getText()          # background or bold
                    val = p.style_val().getText()    # red, true, green etc
                    props[key] = val
                
                self.styles[name] = props
                return

            case QAReporterDSLParser.ApplyContext():
                style_name = t.ID().getText()
                
                if style_name not in self.styles:
                    print(f"[ERROR] Style '{style_name}' not defined.")
                    return

                # Evaluate what changes should be applied to cell
                filter_mask = self.evaluate(t.exp())

                if filter_mask is not None:
                    indexes = self.df[filter_mask].index.tolist()
                    self.filter.append({'indexes': indexes, 'style': style_name})
                    print(f"[APPLY] Style '{style_name}' in {len(indexes)} lines.")
                return

            case QAReporterDSLParser.WriteContext():
                if self.df.empty:
                    print("[WARNING] Nothing to save.")
                    return

                path = t.OUTPUTFILE().getText().strip('"')
                self.output_file = path
                self.df.to_excel(path, index=False)
                
                try:
                    wb = load_workbook(path)
                    ws = wb.active
                except Exception as e:
                    print(f"[ERROR] Failed to open Excel: {e}")
                    return

                for filter in self.filter:
                    style = self.styles.get(filter['style'])
                    if not style: continue

                    fill = None
                    if 'background' in style:
                        color_name = style['background']
                        hex_cor = COLORS_HEX.get(color_name, 'FFFFFF')  # converting color to hexadecimal
                        
                        fill = PatternFill(start_color=hex_cor, end_color=hex_cor, fill_type="solid")  # filling cell on spreadsheet

                    font = Font()
                    if 'bold' in style and style['bold'] == 'true':
                        font = Font(bold=True)

                    for idx in filter['indexes']:
                        row = idx + 2 #
                        for col in range(1, len(self.df.columns) + 1):
                            cell = ws.cell(row=row, column=col)
                            if fill: cell.fill = fill
                            cell.font = font
                
                wb.save(path)
                print(f"[SAVE] File generated: {path}")
                return

            case QAReporterDSLParser.VisualizeContext():
                if not self.output_file:
                    print("[ERROR] Use SAVE before VISUALIZE.")
                    return
            
                view_option = t.view_option().getText()

                if view_option == 'BROWSER':
                    html_file = self.output_file.replace('.xlsx', '.html')
                    xlsx2html(self.output_file, html_file)
                    webbrowser.open(html_file)
                    print(f"Opening HTML visualization for '{self.output_file}' in browser.")

                elif view_option == 'EXCEL': 
                    if os.environ.get('CODESPACES') == 'true':                   # Running in Codespaces (Remote/Linux)
                        print(f"[Codespaces] File '{self.output_file}' generated.")
                        print("To open in Excel, right-click the file in the sidebar and choose 'Download'.")

                    else:                                                        # Running locally (Windows/Linux/MacOS)
                        try:
                            os.startfile(self.output_file)
                            print(f"[Local] Opening '{self.output_file}' in default system application.")
                        except AttributeError: # Linux/MacOs
                            subprocess.call(['xdg-open', self.output_file])
                            print(f"[Local] Opening '{self.output_file}' in default system application.")
                        except FileNotFoundError:
                            print(f"[ERROR] File '{self.output_file}' was not found.")
                return
             

            # Relational expressions (using Pandas notation)

            case QAReporterDSLParser.OrContext():
                left = self.evaluate(t.exp(0))
                right = self.evaluate(t.exp(1))

                if left is None or right is None: # extra protection step to check for missing columns etc
                    return None
                
                return left | right  # in Pandas , or == | 

            case QAReporterDSLParser.AndContext():
                left = self.evaluate(t.exp(0))
                right = self.evaluate(t.exp(1))

                if left is None or right is None: # extra protection step to check for missing columns etc
                    return None
                    
                return left & right  # in Pandas , and == &

            case QAReporterDSLParser.NotContext():
                internal_result = self.evaluate(t.exp())
                
                if internal_result is None: # extra protection step to check for missing columns etc
                    return None
                
                return ~internal_result # in Pandas , not == ~ 

            case QAReporterDSLParser.ParenContext():
                return self.evaluate(t.exp())

            case QAReporterDSLParser.RelationContext():
                col_name = t.ID().getText()
                op = t.OP().getText()
                val_ctx = t.value()
                val = self.eval_value(val_ctx)

                if col_name not in self.df.columns:
                    print(f"[ERROR] Column '{col_name}' not found.")
                    return None

                col = self.df[col_name]
                
                if op == '==': return col == val
                if op == '!=': return col != val
                if op == '>':  return col > val
                if op == '<':  return col < val
                if op == '>=': return col >= val
                if op == '<=': return col <= val

            case QAReporterDSLParser.ContainsContext():
                col_name = t.ID().getText()
                val_ctx = t.value()
                val = self.eval_value(val_ctx)

                if col_name not in self.df.columns:
                    print(f"[ERROR] Column '{col_name}' not found.")
                    return None
                
                col = self.df[col_name]

                # Make sure there are no '"' on val (in case eval_value doesn't work)
                search_value = str(val).strip('"')
                
                # Match substrings
                return col.astype(str).str.contains(search_value, case=False, na=False)

            case _:
                pass

    def eval_value(self, t):
        if t.INT():   return int(t.getText())
        if t.FLOAT(): return float(t.getText())
        if t.BOOL():  return t.getText() == 'true'
        if t.STR():   return t.getText().strip('"')
        return t.getText()